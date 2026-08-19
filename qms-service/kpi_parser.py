
from __future__ import annotations

import math
import os
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from threading import RLock
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


MONTH_NAMES = {
    1: "Januari",
    2: "Februari",
    3: "Maret",
    4: "April",
    5: "Mei",
    6: "Juni",
    7: "Juli",
    8: "Agustus",
    9: "September",
    10: "Oktober",
    11: "November",
    12: "Desember",
}

MONTH_ALIASES = {
    1: {"jan", "januari", "january"},
    2: {"feb", "februari", "february"},
    3: {"mar", "maret", "march"},
    4: {"apr", "april"},
    5: {"mei", "may"},
    6: {"jun", "juni", "june"},
    7: {"jul", "juli", "july"},
    8: {"agu", "agst", "aug", "agustus", "august"},
    9: {"sep", "sept", "september"},
    10: {"okt", "oct", "oktober", "october"},
    11: {"nov", "november"},
    12: {"des", "dec", "desember", "december"},
}

PLAN_LABELS = {"plan", "rencana", "target plan"}
ACTUAL_LABELS = {"actual", "aktual", "realisasi"}

# Penanda bahwa laporan/realisasi belum disampaikan oleh PIC divisi.
# Nilai ini tidak boleh dianggap gagal KPI karena belum ada data yang bisa dinilai.
NO_REPORT_LABELS = {
    "tbd",
    "to be determined",
    "to be defined",
    "belum ada laporan",
    "laporan belum ada",
    "belum ada report",
    "belum dilaporkan",
    "belum laporan",
    "belum setor",
    "belum setoran",
}

LESS_IS_BETTER_WORDS = (
    "maksimal",
    "maximum",
    "max ",
    "tidak melebihi",
    "tidak lebih dari",
    "kurang dari",
    "lebih kecil",
    "di bawah",
    "dibawah",
    "zero incident",
    "zero accident",
    "incident",
    "accident",
    "defect",
    "reject",
    "downtime",
    "complaint",
    "keluhan",
    "budget utilization",
    "realisasi anggaran",

    # KPI yang semakin kecil semakin baik
    "past due",
    "overdue",
    "aging",
    "lead time",
    "completion time",
    "turnaround time",
    "waktu penyelesaian",
    "batas waktu",
    "setelah akhir bulan",
    "selambat-lambatnya",
    "selambat lambatnya",
)
MORE_IS_BETTER_WORDS = (
    "minimal",
    "minimum",
    "min ",
    "sekurang",
    "setidaknya",
    "paling sedikit",
    "lebih dari",
    "lebih besar",
    "di atas",
    "diatas",
)


@dataclass(frozen=True)
class SheetLayout:
    month_header_row: int
    month_columns: dict[int, int]
    plan_actual_column: int
    kpi_column: int | None
    variable_column: int | None
    unit_column: int | None
    target_column: int | None


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip().lower()
    return text


def normalize_label(value: Any) -> str:
    text = normalize_text(value)
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def month_number(value: Any) -> int | None:
    label = normalize_label(value)
    if not label:
        return None

    for number, aliases in MONTH_ALIASES.items():
        if label in aliases:
            return number
    return None


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return normalize_text(value) in {"", "-", "—", "n/a", "na", "none"}
    return False


def is_no_report(value: Any) -> bool:
    """True jika isi cell menandakan laporan belum disampaikan.

    Contoh paling umum pada workbook QMS adalah ``TBD``. Status ini
    dipisahkan dari *Tidak memenuhi* agar direksi tidak membaca data yang
    belum masuk sebagai kegagalan kinerja.
    """
    if value is None:
        return False
    return normalize_label(value) in NO_REPORT_LABELS


def parse_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, (int, float)):
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            return None
        return float(value)

    text = normalize_text(value)
    if not text or text in {"-", "—", "n/a", "na"}:
        return None

    text = text.replace("rp", "").replace("%", "")
    text = text.replace("\u00a0", " ")
    text = re.sub(r"[^0-9,.\-]+", "", text)

    if not text or text in {"-", ".", ","}:
        return None

    negative = text.startswith("-")
    text = text.lstrip("-")

    if "," in text and "." in text:
        last_comma = text.rfind(",")
        last_dot = text.rfind(".")
        decimal_sep = "," if last_comma > last_dot else "."
        thousands_sep = "." if decimal_sep == "," else ","
        text = text.replace(thousands_sep, "")
        text = text.replace(decimal_sep, ".")
    elif "," in text:
        parts = text.split(",")
        if len(parts) > 2 and all(len(part) == 3 for part in parts[1:]):
            text = "".join(parts)
        elif len(parts) == 2 and len(parts[1]) in {1, 2}:
            text = ".".join(parts)
        elif len(parts) == 2 and len(parts[1]) == 3:
            text = "".join(parts)
        else:
            text = "".join(parts)
    elif "." in text:
        parts = text.split(".")
        if len(parts) > 2 and all(len(part) == 3 for part in parts[1:]):
            text = "".join(parts)
        elif len(parts) == 2 and len(parts[1]) in {1, 2}:
            pass
        elif len(parts) == 2 and len(parts[1]) == 3:
            text = "".join(parts)
        elif len(parts) > 2:
            text = "".join(parts)

    try:
        number = float(text)
    except ValueError:
        return None

    return -number if negative else number


def extract_numbers(value: Any) -> list[float]:
    if value is None:
        return []
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return [float(value)]

    text = str(value)
    matches = re.findall(r"-?\d+(?:[.,]\d+)*", text)
    numbers: list[float] = []
    for match in matches:
        number = parse_number(match)
        if number is not None:
            numbers.append(number)
    return numbers


def format_value(value: Any) -> str:
    if value is None:
        return "-"
    if isinstance(value, bool):
        return "Ya" if value else "Tidak"
    if isinstance(value, float):
        if value.is_integer():
            return f"{int(value):,}".replace(",", ".")
        return f"{value:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return str(value).strip() or "-"


def format_cell_value(cell: Any) -> str:
    value = cell.value
    if value is None:
        return "-"

    number_format = str(getattr(cell, "number_format", "") or "")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if "%" in number_format:
            percentage = float(value) * 100
            if percentage.is_integer():
                return f"{int(percentage)}%"
            return f"{percentage:.2f}".replace(".", ",") + "%"

    return format_value(value)


class KpiWorkbookParser:
    def __init__(
        self,
        excel_path: str | os.PathLike[str],
        ignored_sheets: Iterable[str] = ("MENU", "LIST"),
        max_scan_rows: int = 300,
        max_scan_columns: int = 90,
    ) -> None:
        self.excel_path = Path(excel_path)
        self.ignored_sheets = {normalize_text(name) for name in ignored_sheets}
        self.max_scan_rows = max_scan_rows
        self.max_scan_columns = max_scan_columns

        # Cache lookup merged-cell per worksheet. Workbook KPI memiliki banyak
        # merged cell. Versi lama melakukan iterasi ke seluruh merged range
        # setiap kali cell kosong dibaca, sehingga first load dapat memakan
        # puluhan detik. Mapping ini dibuat sekali per worksheet dan lookup
        # berikutnya menjadi O(1).
        self._merged_anchor_cache: dict[
            int, dict[tuple[int, int], tuple[int, int]]
        ] = {}

    def _merged_anchor_map(
        self,
        worksheet: Worksheet,
    ) -> dict[tuple[int, int], tuple[int, int]]:
        """Mengembalikan map posisi merged-cell -> cell kiri-atas.

        Hanya area yang mungkin dipindai parser yang dimasukkan ke cache agar
        penggunaan memori tetap kecil meski worksheet memiliki formatting
        sampai ribuan baris.
        """
        worksheet_key = id(worksheet)
        cached = self._merged_anchor_cache.get(worksheet_key)
        if cached is not None:
            return cached

        mapping: dict[tuple[int, int], tuple[int, int]] = {}
        scan_row_limit = min(worksheet.max_row, self.max_scan_rows)
        scan_column_limit = min(worksheet.max_column, self.max_scan_columns)

        for merged_range in worksheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds

            # Range tidak bersinggungan dengan area parser.
            if min_row > scan_row_limit or min_col > scan_column_limit:
                continue
            if max_row < 1 or max_col < 1:
                continue

            anchor = (min_row, min_col)
            clipped_max_row = min(max_row, scan_row_limit)
            clipped_max_col = min(max_col, scan_column_limit)

            for row in range(max(1, min_row), clipped_max_row + 1):
                for column in range(max(1, min_col), clipped_max_col + 1):
                    mapping[(row, column)] = anchor

        self._merged_anchor_cache[worksheet_key] = mapping
        return mapping

    def _effective_cell_value(
        self,
        worksheet: Worksheet,
        row: int,
        column: int,
    ) -> Any:
        """Mengambil nilai cell, termasuk nilai induk dari merged cell.

        Beberapa sheet seperti LEGAL memakai merge vertikal untuk nama KPI.
        OpenPyXL hanya menyimpan nilai pada cell kiri-atas; cell lain di area
        merge terlihat kosong. Helper ini membuat pembacaan tetap konsisten.
        """
        cell = worksheet.cell(row, column)
        if cell.value is not None:
            return cell.value

        anchor = self._merged_anchor_map(worksheet).get((row, column))
        if anchor is None:
            return None

        anchor_row, anchor_column = anchor
        return worksheet.cell(anchor_row, anchor_column).value

    def parse(self) -> dict[str, Any]:
        # Workbook baru berarti object worksheet baru; buang cache merged-cell
        # dari parse sebelumnya agar tidak pernah memakai referensi stale.
        self._merged_anchor_cache.clear()

        if not self.excel_path.is_file():
            raise FileNotFoundError(
                f"File Excel tidak ditemukan: {self.excel_path}"
            )

        workbook = load_workbook(
            filename=self.excel_path,
            data_only=True,
            read_only=False,
        )

        try:
            monthly_divisions: dict[int, list[dict[str, Any]]] = {
                month: [] for month in MONTH_NAMES
            }
            diagnostics: list[dict[str, Any]] = []

            for worksheet in workbook.worksheets:
                if normalize_text(worksheet.title) in self.ignored_sheets:
                    continue

                layout = self._detect_layout(worksheet)
                if layout is None:
                    diagnostics.append(
                        {
                            "division": worksheet.title,
                            "status": "layout_tidak_terdeteksi",
                            "message": (
                                "Header bulan atau pasangan Plan/Actual "
                                "tidak ditemukan."
                            ),
                        }
                    )
                    continue

                for month in MONTH_NAMES:
                    result = self._analyze_division_month(
                        worksheet=worksheet,
                        layout=layout,
                        month=month,
                    )
                    monthly_divisions[month].append(result)

            months = {
                str(month): self._build_month_summary(month, divisions)
                for month, divisions in monthly_divisions.items()
            }

            modified_at = datetime.fromtimestamp(
                self.excel_path.stat().st_mtime
            ).isoformat(timespec="seconds")

            return {
                "source": {
                    "file_name": self.excel_path.name,
                    "file_path": str(self.excel_path),
                    "modified_at": modified_at,
                },
                "months": months,
                "diagnostics": diagnostics,
            }
        finally:
            workbook.close()

    def _detect_layout(self, worksheet: Worksheet) -> SheetLayout | None:
        max_row = min(worksheet.max_row, self.max_scan_rows)
        max_column = min(worksheet.max_column, self.max_scan_columns)

        month_header_row = 0
        month_columns: dict[int, int] = {}

        for row in range(1, max_row + 1):
            candidate: dict[int, int] = {}
            for column in range(1, max_column + 1):
                month = month_number(worksheet.cell(row, column).value)
                if month is not None:
                    candidate[month] = column

            if len(candidate) > len(month_columns):
                month_header_row = row
                month_columns = candidate

        if len(month_columns) < 3:
            return None

        plan_actual_column = self._detect_plan_actual_column(
            worksheet,
            start_row=month_header_row + 1,
            max_row=max_row,
            max_column=max_column,
            month_columns=month_columns,
        )
        if plan_actual_column is None:
            return None

        # Pada beberapa template (terutama LEGAL), header Jan berada satu
        # kolom sebelum data Jan karena kolom tersebut masih berisi label
        # Plan/Actual. Deteksi dan koreksi offset ini secara otomatis.
        month_data_shift = self._detect_month_data_shift(
            worksheet=worksheet,
            month_columns=month_columns,
            plan_actual_column=plan_actual_column,
            start_row=month_header_row + 1,
            max_row=max_row,
            max_column=max_column,
        )
        if month_data_shift:
            month_columns = {
                month: column + month_data_shift
                for month, column in month_columns.items()
                if column + month_data_shift <= max_column
            }

        field_columns = self._detect_field_columns(
            worksheet,
            month_header_row=month_header_row,
            max_column=max_column,
        )

        kpi_column = field_columns.get("kpi")
        if kpi_column is None:
            kpi_column = self._fallback_kpi_column(
                worksheet,
                month_header_row=month_header_row,
                plan_actual_column=plan_actual_column,
                max_row=max_row,
            )

        return SheetLayout(
            month_header_row=month_header_row,
            month_columns=month_columns,
            plan_actual_column=plan_actual_column,
            kpi_column=kpi_column,
            variable_column=field_columns.get("variable"),
            unit_column=field_columns.get("unit"),
            target_column=field_columns.get("target"),
        )

    def _detect_plan_actual_column(
        self,
        worksheet: Worksheet,
        start_row: int,
        max_row: int,
        max_column: int,
        month_columns: dict[int, int],
    ) -> int | None:
        """Mencari kolom Plan/Actual yang terkait dengan tabel bulan.

        Versi lama menilai semua kolom di seluruh sheet. Jika ada tabel lain
        di bagian bawah, kolom Plan/Actual dari tabel lain dapat terpilih dan
        membuat BusDev terbaca kosong. Sekarang kolom di sebelah kiri bulan
        pertama diprioritaskan, sesuai struktur workbook QMS.
        """
        first_month_column = min(month_columns.values())
        local_start = max(1, first_month_column - 4)
        local_end = max(local_start, first_month_column - 1)

        def score_column(column: int) -> int:
            plan_count = 0
            actual_count = 0
            # Batasi penilaian ke area tabel terdekat agar tabel lain pada
            # sheet yang sama tidak mengambil alih deteksi.
            scan_end = min(max_row, start_row + 180)
            for row in range(start_row, scan_end + 1):
                label = normalize_label(
                    self._effective_cell_value(worksheet, row, column)
                )
                if label in PLAN_LABELS:
                    plan_count += 1
                elif label in ACTUAL_LABELS:
                    actual_count += 1
            proximity_bonus = max(0, 5 - (first_month_column - column))
            return min(plan_count, actual_count) * 20 + plan_count + actual_count + proximity_bonus

        best_column: int | None = None
        best_score = 0

        for column in range(local_start, local_end + 1):
            score = score_column(column)
            if score > best_score:
                best_score = score
                best_column = column

        if best_column is not None and best_score >= 22:
            return best_column

        # Fallback untuk template yang benar-benar berbeda.
        for column in range(1, max_column + 1):
            score = score_column(column)
            if score > best_score:
                best_score = score
                best_column = column

        return best_column if best_score >= 22 else None

    def _detect_month_data_shift(
        self,
        worksheet: Worksheet,
        month_columns: dict[int, int],
        plan_actual_column: int,
        start_row: int,
        max_row: int,
        max_column: int,
    ) -> int:
        """Mengoreksi header bulan yang satu kolom lebih kiri dari data.

        Contoh sheet LEGAL:
        - kolom R berisi pasangan Plan/Actual utama;
        - header Jan berada di S;
        - tetapi S masih berisi teks Plan/Actual;
        - nilai Januari sebenarnya mulai di T.
        """
        if not month_columns:
            return 0

        first_month_column = month_columns[min(month_columns)]
        if first_month_column + 1 > max_column:
            return 0

        shifted_label_pairs = 0
        normal_data_pairs = 0
        row = start_row
        scan_end = min(max_row, start_row + 180)

        while row <= scan_end:
            label = normalize_label(
                self._effective_cell_value(worksheet, row, plan_actual_column)
            )
            if label not in PLAN_LABELS:
                row += 1
                continue

            actual_row = self._find_actual_row(
                worksheet,
                plan_row=row,
                plan_actual_column=plan_actual_column,
                max_row=scan_end,
            )
            if actual_row is None:
                row += 1
                continue

            plan_at_header = normalize_label(
                self._effective_cell_value(worksheet, row, first_month_column)
            )
            actual_at_header = normalize_label(
                self._effective_cell_value(worksheet, actual_row, first_month_column)
            )

            if plan_at_header in PLAN_LABELS and actual_at_header in ACTUAL_LABELS:
                shifted_label_pairs += 1
            elif not (is_blank(plan_at_header) and is_blank(actual_at_header)):
                normal_data_pairs += 1

            row = actual_row + 1

        if shifted_label_pairs > 0 and shifted_label_pairs >= normal_data_pairs:
            return 1
        return 0

    def _detect_field_columns(
        self,
        worksheet: Worksheet,
        month_header_row: int,
        max_column: int,
    ) -> dict[str, int]:
        fields: dict[str, int] = {}
        start_row = max(1, month_header_row - 5)

        keyword_map = {
            "kpi": (
                "key performance indicator",
                "indikator kinerja utama",
            ),
            "variable": ("variable", "variabel"),
            "unit": (
                "unit of measurement",
                "satuan pengukuran",
                "unit measurement",
            ),
            "target": ("target", "sasaran"),
        }

        for column in range(1, max_column + 1):
            combined = " ".join(
                normalize_text(self._effective_cell_value(worksheet, row, column))
                for row in range(start_row, month_header_row + 1)
            )

            for field, keywords in keyword_map.items():
                if field in fields:
                    continue
                if any(keyword in combined for keyword in keywords):
                    fields[field] = column

        return fields

    def _fallback_kpi_column(
        self,
        worksheet: Worksheet,
        month_header_row: int,
        plan_actual_column: int,
        max_row: int,
    ) -> int | None:
        best_column: int | None = None
        best_score = 0

        for column in range(1, plan_actual_column):
            score = 0
            for row in range(month_header_row + 1, max_row + 1):
                value = self._effective_cell_value(worksheet, row, column)
                text = normalize_text(value)
                if text and not text.isdigit() and len(text) >= 3:
                    score += min(len(text), 50)

            if score > best_score:
                best_score = score
                best_column = column

        return best_column

    def _analyze_division_month(
        self,
        worksheet: Worksheet,
        layout: SheetLayout,
        month: int,
    ) -> dict[str, Any]:
        month_column = layout.month_columns.get(month)
        if month_column is None:
            return self._division_result(
                worksheet.title,
                month,
                [],
                message="Kolom bulan tidak ditemukan pada sheet.",
            )

        max_row = min(worksheet.max_row, self.max_scan_rows)
        items: list[dict[str, Any]] = []

        row = layout.month_header_row + 1
        while row <= max_row:
            label = normalize_label(
                self._effective_cell_value(
                    worksheet, row, layout.plan_actual_column
                )
            )

            if label not in PLAN_LABELS:
                row += 1
                continue

            actual_row = self._find_actual_row(
                worksheet,
                plan_row=row,
                plan_actual_column=layout.plan_actual_column,
                max_row=max_row,
            )
            if actual_row is None:
                row += 1
                continue

            kpi_name = self._nearby_value(
                worksheet,
                row,
                layout.kpi_column,
                allowed_rows=(0, -1, -2, -3, -4, -5, -6),
            )
            variable = self._nearby_value(
                worksheet,
                row,
                layout.variable_column,
                allowed_rows=(0, -1, -2, -3, -4, -5, -6),
            )
            unit = self._nearby_value(
                worksheet,
                row,
                layout.unit_column,
                allowed_rows=(0, -1, -2, -3, -4, -5, -6),
            )
            target = self._nearby_value(
                worksheet,
                row,
                layout.target_column,
                allowed_rows=(0, -1, -2, -3, -4, -5, -6),
            )

            plan_cell = worksheet.cell(row, month_column)
            actual_cell = worksheet.cell(actual_row, month_column)
            plan_value = plan_cell.value
            actual_value = actual_cell.value

            # Abaikan pasangan Plan/Actual di luar tabel KPI yang tidak
            # mempunyai konteks indikator sama sekali. Ini mencegah baris
            # bantu/footer terbaca sebagai KPI kosong dan membuat satu divisi
            # keliru masuk kategori belum dinilai.
            if all(
                is_blank(value)
                for value in (kpi_name, variable, unit, target)
            ):
                row = actual_row + 1
                continue

            if is_blank(kpi_name):
                kpi_name = f"KPI baris {row}"

            evaluation = self._evaluate(
                kpi_name=str(kpi_name),
                variable=variable,
                unit=unit,
                target=target,
                plan=plan_value,
                actual=actual_value,
            )

            items.append(
                {
                    "kpi": str(kpi_name),
                    "variable": format_value(variable),
                    "unit": format_value(unit),
                    "target": format_value(target),
                    "plan": format_cell_value(plan_cell),
                    "actual": format_cell_value(actual_cell),
                    "status": evaluation["status"],
                    "status_label": evaluation["status_label"],
                    "reason": evaluation["reason"],
                    "plan_row": row,
                    "actual_row": actual_row,
                }
            )

            row = actual_row + 1

        return self._division_result(worksheet.title, month, items)

    def _find_actual_row(
        self,
        worksheet: Worksheet,
        plan_row: int,
        plan_actual_column: int,
        max_row: int,
    ) -> int | None:
        for row in range(plan_row + 1, min(plan_row + 4, max_row) + 1):
            label = normalize_label(
                self._effective_cell_value(
                    worksheet, row, plan_actual_column
                )
            )
            if label in ACTUAL_LABELS:
                return row
            if label in PLAN_LABELS:
                return None
        return None

    def _nearby_value(
        self,
        worksheet: Worksheet,
        base_row: int,
        column: int | None,
        allowed_rows: tuple[int, ...],
    ) -> Any:
        if column is None:
            return None

        for offset in allowed_rows:
            row = base_row + offset
            if row < 1:
                continue
            value = self._effective_cell_value(worksheet, row, column)
            if not is_blank(value):
                return value
        return None

    def _evaluate(
        self,
        kpi_name: str,
        variable: Any,
        unit: Any,
        target: Any,
        plan: Any,
        actual: Any,
    ) -> dict[str, str]:
        if is_blank(plan) and is_blank(actual):
            return {
                "status": "tidak_dijadwalkan",
                "status_label": "Tidak dijadwalkan",
                "reason": "Plan dan Actual bulan ini kosong.",
            }

        if is_blank(plan):
            return {
                "status": "belum_lengkap",
                "status_label": "Belum lengkap",
                "reason": "Actual tersedia, tetapi Plan bulan ini kosong.",
            }

        # TBD berarti laporan/realisasi belum diterima. Ini bukan kegagalan KPI.
        if is_no_report(actual):
            return {
                "status": "belum_ada_laporan",
                "status_label": "Laporan belum diterima",
                "reason": (
                    "Nilai Actual ditulis TBD, sehingga belum dinilai sebagai "
                    "memenuhi atau tidak memenuhi."
                ),
            }

        if is_blank(actual):
            return {
                "status": "belum_lengkap",
                "status_label": "Belum lengkap",
                "reason": "Plan tersedia, tetapi Actual bulan ini belum diisi.",
            }

        plan_number = parse_number(plan)
        actual_number = parse_number(actual)

        context = " ".join(
            normalize_text(value)
            for value in (kpi_name, variable, unit, target, plan)
            if value is not None
        )

        comparator = self._infer_comparator(context)
        plan_numbers = extract_numbers(plan)

        # KPI berbentuk teks tetap dapat dinilai. Versi sebelumnya langsung
        # menganggap Actual non-angka sebagai data belum lengkap, walaupun Plan
        # dan Actual sama-sama terisi. Hal ini dapat membuat sheet seperti
        # BusDev keliru masuk kategori belum dinilai.
        threshold = plan_number
        if threshold is None and plan_numbers:
            threshold = plan_numbers[-1]

        if threshold is None:
            plan_text = normalize_label(plan)
            actual_text = normalize_label(actual)
            passed = bool(plan_text) and plan_text == actual_text
            return self._comparison_result(
                passed,
                "Plan dan Actual dibandingkan sebagai teks karena keduanya bukan angka.",
            )

        if actual_number is None:
            return {
                "status": "belum_lengkap",
                "status_label": "Belum lengkap",
                "reason": (
                    "Plan berupa angka, tetapi nilai Actual tidak dapat dibaca "
                    "sebagai angka."
                ),
            }

        if comparator == "range" and len(plan_numbers) >= 2:
            lower, upper = min(plan_numbers), max(plan_numbers)
            passed = lower <= actual_number <= upper
            return self._comparison_result(
                passed,
                f"Actual harus berada di antara {lower:g} dan {upper:g}.",
            )

        tolerance = max(abs(threshold), 1.0) * 1e-9

        if comparator == "<=":
            passed = actual_number <= threshold + tolerance
            reason = (
                f"Actual harus lebih kecil atau sama dengan Plan "
                f"({actual_number:g} ≤ {threshold:g})."
            )
        elif comparator == "<":
            passed = actual_number < threshold - tolerance
            reason = (
                f"Actual harus lebih kecil dari Plan "
                f"({actual_number:g} < {threshold:g})."
            )
        elif comparator == ">":
            passed = actual_number > threshold + tolerance
            reason = (
                f"Actual harus lebih besar dari Plan "
                f"({actual_number:g} > {threshold:g})."
            )
        else:
            passed = actual_number + tolerance >= threshold
            reason = (
                f"Actual harus lebih besar atau sama dengan Plan "
                f"({actual_number:g} ≥ {threshold:g})."
            )

        return self._comparison_result(passed, reason)

    def _infer_comparator(self, context: Any) -> str:
        """Menentukan arah evaluasi KPI dari teks target/plan.

        Nilai bawaan adalah ``>=`` karena sebagian besar KPI menilai bahwa
        realisasi yang lebih tinggi dari rencana berarti memenuhi. KPI seperti
        biaya, keluhan, keterlambatan, insiden, dan lead time menggunakan arah
        sebaliknya apabila konteksnya mengandung penanda maksimum/lebih kecil.
        """
        text = normalize_text(context)
        if not text:
            return ">="

        numbers = extract_numbers(text)
        range_words = (
            "antara",
            "rentang",
            "range",
            "sampai dengan",
            "s.d.",
            " s/d ",
        )
        has_dash_range = bool(
            re.search(r"\d\s*(?:-|–|—)\s*\d", text)
        )
        if len(numbers) >= 2 and (
            any(word in text for word in range_words) or has_dash_range
        ):
            return "range"

        # Simbol eksplisit selalu lebih kuat daripada kata kunci umum.
        if any(symbol in text for symbol in ("<=", "=<", "≤")):
            return "<="
        if any(symbol in text for symbol in (">=", "=>", "≥")):
            return ">="
        if re.search(r"(?<![<>=])<(?![=])", text):
            return "<"
        if re.search(r"(?<![<>=])>(?![=])", text):
            return ">"

        # Periksa kata 'lebih kecil' terlebih dahulu karena beberapa frasa
        # mengandung kata 'lebih dari' di dalam kalimat negatif.
        if any(word in text for word in LESS_IS_BETTER_WORDS):
            return "<="
        if any(word in text for word in MORE_IS_BETTER_WORDS):
            return ">="

        return ">="

    def _comparison_result(self, passed: bool, reason: str) -> dict[str, str]:
        if passed:
            return {
                "status": "memenuhi",
                "status_label": "Memenuhi",
                "reason": reason,
            }
        return {
            "status": "tidak_memenuhi",
            "status_label": "Tidak memenuhi",
            "reason": reason,
        }

    def _division_result(
        self,
        division: str,
        month: int,
        items: list[dict[str, Any]],
        message: str | None = None,
    ) -> dict[str, Any]:
        counts = {
            "memenuhi": 0,
            "tidak_memenuhi": 0,
            "belum_ada_laporan": 0,
            "belum_lengkap": 0,
            "tidak_dijadwalkan": 0,
        }

        for item in items:
            status = item["status"]
            if status in counts:
                counts[status] += 1

        # Jika masih ada Actual bertanda TBD/kosong, status divisi belum
        # boleh diputuskan gagal. PIC belum menyerahkan seluruh laporan bulan
        # tersebut. Detail KPI yang sudah gagal tetap terlihat di popup, tetapi
        # ringkasan divisi masuk kategori laporan belum diterima.
        if counts["belum_ada_laporan"] > 0:
            division_status = "belum_ada_laporan"
            division_status_label = "Laporan belum diterima"
        elif counts["belum_lengkap"] > 0:
            division_status = "belum_lengkap"
            division_status_label = "Laporan belum lengkap"
        elif counts["tidak_memenuhi"] > 0:
            division_status = "tidak_memenuhi"
            division_status_label = "Tidak memenuhi"
        elif counts["memenuhi"] > 0:
            division_status = "memenuhi"
            division_status_label = "Memenuhi"
        else:
            division_status = "belum_ada_data"
            division_status_label = "Belum ada data"

        evaluated = counts["memenuhi"] + counts["tidak_memenuhi"]
        achievement = (
            round((counts["memenuhi"] / evaluated) * 100, 2)
            if evaluated
            else 0.0
        )

        return {
            "division": division.strip(),
            "month": month,
            "month_name": MONTH_NAMES[month],
            "status": division_status,
            "status_label": division_status_label,
            "achievement_percentage": achievement,
            "counts": counts,
            "items": items,
            "message": message,
        }

    def _build_month_summary(
        self,
        month: int,
        divisions: list[dict[str, Any]],
    ) -> dict[str, Any]:
        order = {
            "tidak_memenuhi": 0,
            "belum_ada_laporan": 1,
            "belum_lengkap": 2,
            "memenuhi": 3,
            "belum_ada_data": 4,
        }
        divisions = sorted(
            divisions,
            key=lambda item: (
                order.get(item["status"], 9),
                item["division"].lower(),
            ),
        )

        meeting = [
            item["division"]
            for item in divisions
            if item["status"] == "memenuhi"
        ]
        not_meeting = [
            item["division"]
            for item in divisions
            if item["status"] == "tidak_memenuhi"
        ]
        no_report = [
            item["division"]
            for item in divisions
            if item["status"] == "belum_ada_laporan"
        ]
        data_incomplete = [
            item["division"]
            for item in divisions
            if item["status"] in {"belum_lengkap", "belum_ada_data"}
        ]
        pending = no_report + data_incomplete

        assessed = len(meeting) + len(not_meeting)
        compliance = round((len(meeting) / assessed) * 100, 2) if assessed else 0.0

        return {
            "month": month,
            "month_name": MONTH_NAMES[month],
            "summary": {
                "total_divisions": len(divisions),
                "assessed_divisions": assessed,
                "meeting_count": len(meeting),
                "not_meeting_count": len(not_meeting),
                # Tetap mempertahankan nama lama agar UI versi sebelumnya
                # tidak rusak, tetapi nilainya adalah seluruh data pending.
                "incomplete_count": len(pending),
                "report_pending_count": len(pending),
                "no_report_count": len(no_report),
                "data_incomplete_count": len(data_incomplete),
                "compliance_percentage": compliance,
            },
            "meeting_divisions": meeting,
            "not_meeting_divisions": not_meeting,
            "no_report_divisions": no_report,
            "incomplete_divisions": pending,
            "data_incomplete_divisions": data_incomplete,
            "divisions": divisions,
        }


class CachedKpiRepository:
    def __init__(self, parser: KpiWorkbookParser) -> None:
        self.parser = parser
        self._lock = RLock()
        self._cached_mtime_ns: int | None = None
        self._cached_data: dict[str, Any] | None = None

    def set_excel_path(self, excel_path: str | os.PathLike[str]) -> None:
        """Mengganti workbook aktif dan mengosongkan cache secara aman."""
        with self._lock:
            self.parser.excel_path = Path(excel_path)
            self._cached_mtime_ns = None
            self._cached_data = None

    def clear(self) -> None:
        with self._lock:
            self._cached_mtime_ns = None
            self._cached_data = None

    def get(self, force_refresh: bool = False) -> dict[str, Any]:
        path = self.parser.excel_path
        if not path.is_file():
            raise FileNotFoundError(f"File Excel tidak ditemukan: {path}")

        mtime_ns = path.stat().st_mtime_ns

        with self._lock:
            if (
                force_refresh
                or self._cached_data is None
                or self._cached_mtime_ns != mtime_ns
            ):
                self._cached_data = self.parser.parse()
                self._cached_mtime_ns = mtime_ns

            return self._cached_data

from __future__ import annotations

import os
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from threading import RLock
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


RISK_TYPE_ORDER = ["Strategic", "Operational", "Financial", "Compliance"]
RISK_TYPE_ALIASES = {
    "strategic": "Strategic",
    "strategy": "Strategic",
    "operational": "Operational",
    "operation": "Operational",
    "financial": "Financial",
    "finance": "Financial",
    "compliance": "Compliance",
    "legal compliance": "Compliance",
}

GRADE_ORDER = ["A", "B", "C", "D", "E"]
GRADE_LABELS = {
    "A": "Rendah",
    "B": "Cukup Rendah",
    "C": "Sedang",
    "D": "Tinggi",
    "E": "Sangat Tinggi",
}
GRADE_RANK = {grade: index for index, grade in enumerate(GRADE_ORDER)}


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\u00a0", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_risk_type(value: Any) -> str | None:
    key = normalize_key(value)
    if not key:
        return None
    if key in RISK_TYPE_ALIASES:
        return RISK_TYPE_ALIASES[key]
    for alias, canonical in RISK_TYPE_ALIASES.items():
        if alias in key:
            return canonical
    return normalize_text(value).title()


def normalize_grade(value: Any) -> str | None:
    text = normalize_text(value).upper()
    if not text:
        return None
    match = re.search(r"\b([A-E])\b", text)
    return match.group(1) if match else None


def percentage(part: int, total: int) -> float:
    if total <= 0:
        return 0.0
    return round((part / total) * 100, 1)


def format_cell_value(value: Any) -> str:
    if value is None:
        return "-"
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, bool):
        return "Ya" if value else "Tidak"
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return f"{value:g}"
    text = normalize_text(value)
    return text or "-"


def numeric_cell_value(value: Any) -> float | int | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return int(number) if number.is_integer() else round(number, 2)
    match = re.search(r"-?\d+(?:[.,]\d+)?", normalize_text(value))
    if not match:
        return None
    try:
        number = float(match.group(0).replace(",", "."))
    except ValueError:
        return None
    return int(number) if number.is_integer() else round(number, 2)


class RiskWorkbookParser:
    """Membaca ringkasan dan detail Risk Assessment dari workbook konsolidasi.

    Dashboard memakai tabel detail sebagai sumber utama. Dengan begitu grafik,
    filter, dan daftar risiko tetap akurat walaupun formula pada sheet ringkasan
    workbook belum dihitung ulang.
    """

    def __init__(
        self,
        excel_path: str | os.PathLike[str],
        sheet_name: str = "Risk Assessment",
        max_scan_rows: int = 5000,
        max_scan_columns: int = 80,
    ) -> None:
        self.excel_path = Path(excel_path)
        self.sheet_name = sheet_name
        self.max_scan_rows = max_scan_rows
        self.max_scan_columns = max_scan_columns

    def _find_sheet(self, workbook) -> Worksheet:
        expected = normalize_key(self.sheet_name)
        for worksheet in workbook.worksheets:
            if normalize_key(worksheet.title) == expected:
                return worksheet
        for worksheet in workbook.worksheets:
            title_key = normalize_key(worksheet.title)
            if "risk" in title_key and "assessment" in title_key:
                return worksheet
        raise ValueError(
            f"Sheet Risk Assessment tidak ditemukan pada {self.excel_path.name}."
        )

    def _find_header(self, worksheet: Worksheet) -> tuple[int, dict[str, int]]:
        direct_aliases = {
            "number": {"no", "nomor"},
            "risk_type": {"risk type", "tipe risiko", "kategori risiko"},
            "division": {
                "process division",
                "division",
                "divisi department",
                "process",
                "department",
                "divisi",
            },
            "risk_id": {"risk id", "id risiko"},
            "description": {
                "description of risk",
                "risk description",
                "deskripsi risiko",
            },
            "mitigation": {
                "mitigation countermeasure",
                "mitigation",
                "countermeasure",
                "mitigasi",
            },
            "pic": {"pic", "person in charge", "owner"},
            "due_date": {"due date", "target date", "tanggal target"},
        }

        max_row = min(worksheet.max_row, self.max_scan_rows)
        max_column = min(worksheet.max_column, self.max_scan_columns)

        for row_number in range(1, max_row + 1):
            mapped: dict[str, int] = {}
            score_columns: list[int] = []
            grade_columns: list[int] = []

            for column_number in range(1, max_column + 1):
                key = normalize_key(worksheet.cell(row_number, column_number).value)
                if not key:
                    continue

                for field, aliases in direct_aliases.items():
                    if key in aliases:
                        mapped.setdefault(field, column_number)

                if "risk score" in key:
                    score_columns.append(column_number)
                if "grade" in key and ("risk" in key or "residual" in key or "risidual" in key):
                    grade_columns.append(column_number)

            if score_columns:
                mapped["initial_score"] = score_columns[0]
                if len(score_columns) > 1:
                    mapped["residual_score"] = score_columns[-1]
            if grade_columns:
                mapped["initial_grade"] = grade_columns[0]
                if len(grade_columns) > 1:
                    mapped["residual_grade"] = grade_columns[-1]

            required = {"risk_type", "initial_grade", "residual_grade"}
            if required.issubset(mapped):
                return row_number, mapped

        raise ValueError(
            "Header Risk Type, Risk Grade, dan Residual Grade tidak terdeteksi."
        )

    @staticmethod
    def _cell(worksheet: Worksheet, row: int, columns: dict[str, int], field: str) -> Any:
        column = columns.get(field)
        return worksheet.cell(row, column).value if column else None

    def parse(self) -> dict[str, Any]:
        if not self.excel_path.is_file():
            raise FileNotFoundError(
                f"File Risk Assessment tidak ditemukan: {self.excel_path}"
            )

        workbook = load_workbook(
            filename=self.excel_path,
            data_only=True,
            read_only=False,
        )

        try:
            worksheet = self._find_sheet(workbook)
            header_row, columns = self._find_header(worksheet)

            risk_type_counts: Counter[str] = Counter()
            initial_grade_counts: Counter[str] = Counter()
            residual_grade_counts: Counter[str] = Counter()
            division_counts: Counter[str] = Counter()
            division_high_before: Counter[str] = Counter()
            division_high_after: Counter[str] = Counter()
            type_high_before: Counter[str] = Counter()
            type_high_after: Counter[str] = Counter()
            transition_counts: Counter[str] = Counter()
            records: list[dict[str, Any]] = []
            skipped_rows = 0

            max_row = min(worksheet.max_row, self.max_scan_rows)
            for row_number in range(header_row + 1, max_row + 1):
                risk_type = normalize_risk_type(
                    self._cell(worksheet, row_number, columns, "risk_type")
                )
                initial_grade = normalize_grade(
                    self._cell(worksheet, row_number, columns, "initial_grade")
                )
                residual_grade = normalize_grade(
                    self._cell(worksheet, row_number, columns, "residual_grade")
                )

                if risk_type is None and initial_grade is None and residual_grade is None:
                    continue
                if risk_type is None or initial_grade is None or residual_grade is None:
                    skipped_rows += 1
                    continue

                division = format_cell_value(
                    self._cell(worksheet, row_number, columns, "division")
                )
                risk_id = format_cell_value(
                    self._cell(worksheet, row_number, columns, "risk_id")
                )
                description = format_cell_value(
                    self._cell(worksheet, row_number, columns, "description")
                )
                mitigation = format_cell_value(
                    self._cell(worksheet, row_number, columns, "mitigation")
                )
                pic = format_cell_value(
                    self._cell(worksheet, row_number, columns, "pic")
                )
                due_date = format_cell_value(
                    self._cell(worksheet, row_number, columns, "due_date")
                )
                initial_score = numeric_cell_value(
                    self._cell(worksheet, row_number, columns, "initial_score")
                )
                residual_score = numeric_cell_value(
                    self._cell(worksheet, row_number, columns, "residual_score")
                )

                risk_type_counts[risk_type] += 1
                initial_grade_counts[initial_grade] += 1
                residual_grade_counts[residual_grade] += 1

                if division != "-":
                    division_counts[division] += 1
                if initial_grade in {"D", "E"}:
                    division_high_before[division] += 1
                    type_high_before[risk_type] += 1
                if residual_grade in {"D", "E"}:
                    division_high_after[division] += 1
                    type_high_after[risk_type] += 1

                before_rank = GRADE_RANK[initial_grade]
                after_rank = GRADE_RANK[residual_grade]
                if after_rank < before_rank:
                    transition = "improved"
                elif after_rank > before_rank:
                    transition = "worsened"
                else:
                    transition = "unchanged"
                transition_counts[transition] += 1

                records.append(
                    {
                        "row_number": row_number,
                        "number": format_cell_value(
                            self._cell(worksheet, row_number, columns, "number")
                        ),
                        "risk_type": risk_type,
                        "division": division,
                        "risk_id": risk_id,
                        "description": description,
                        "before_score": initial_score,
                        "before_grade": initial_grade,
                        "before_level": GRADE_LABELS[initial_grade],
                        "after_score": residual_score,
                        "after_grade": residual_grade,
                        "after_level": GRADE_LABELS[residual_grade],
                        "mitigation": mitigation,
                        "pic": pic,
                        "due_date": due_date,
                        "transition": transition,
                    }
                )

            total_risks = len(records)
            if total_risks < 1:
                raise ValueError("Tidak ada baris risiko yang berhasil dibaca.")

            known_types = [
                {
                    "key": normalize_key(risk_type).replace(" ", "_"),
                    "label": risk_type,
                    "count": int(risk_type_counts.get(risk_type, 0)),
                    "percentage": percentage(
                        int(risk_type_counts.get(risk_type, 0)), total_risks
                    ),
                    "high_before": int(type_high_before.get(risk_type, 0)),
                    "high_after": int(type_high_after.get(risk_type, 0)),
                }
                for risk_type in RISK_TYPE_ORDER
            ]
            extra_types = [
                {
                    "key": normalize_key(risk_type).replace(" ", "_"),
                    "label": risk_type,
                    "count": int(count),
                    "percentage": percentage(int(count), total_risks),
                    "high_before": int(type_high_before.get(risk_type, 0)),
                    "high_after": int(type_high_after.get(risk_type, 0)),
                }
                for risk_type, count in sorted(risk_type_counts.items())
                if risk_type not in RISK_TYPE_ORDER
            ]

            grade_comparison = []
            for grade in GRADE_ORDER:
                before = int(initial_grade_counts.get(grade, 0))
                after = int(residual_grade_counts.get(grade, 0))
                grade_comparison.append(
                    {
                        "grade": grade,
                        "level": GRADE_LABELS[grade],
                        "before": before,
                        "after": after,
                        "difference": after - before,
                    }
                )

            high_risk_before = sum(initial_grade_counts.get(grade, 0) for grade in ("D", "E"))
            high_risk_after = sum(residual_grade_counts.get(grade, 0) for grade in ("D", "E"))
            high_risk_reduction = max(0, high_risk_before - high_risk_after)

            division_breakdown = [
                {
                    "division": division,
                    "count": int(count),
                    "high_before": int(division_high_before.get(division, 0)),
                    "high_after": int(division_high_after.get(division, 0)),
                }
                for division, count in sorted(
                    division_counts.items(),
                    key=lambda item: (-item[1], item[0].casefold()),
                )
            ]

            modified_at = datetime.fromtimestamp(
                self.excel_path.stat().st_mtime
            ).isoformat(timespec="seconds")

            return {
                "available": True,
                "total_risks": total_risks,
                "total_divisions": len(division_counts),
                "risk_types": known_types + extra_types,
                "grade_comparison": grade_comparison,
                "high_risk": {
                    "before": int(high_risk_before),
                    "after": int(high_risk_after),
                    "reduction": int(high_risk_reduction),
                    "reduction_percentage": percentage(
                        int(high_risk_reduction), int(high_risk_before)
                    ),
                },
                "mitigation_effectiveness": {
                    "improved": int(transition_counts.get("improved", 0)),
                    "unchanged": int(transition_counts.get("unchanged", 0)),
                    "worsened": int(transition_counts.get("worsened", 0)),
                },
                "division_breakdown": division_breakdown,
                "records": records,
                "source": {
                    "file_name": self.excel_path.name,
                    "sheet_name": worksheet.title,
                    "modified_at": modified_at,
                    "rows_read": total_risks,
                    "skipped_rows": skipped_rows,
                },
            }
        finally:
            workbook.close()


class CachedRiskRepository:
    """Cache ringan berdasarkan path, ukuran file, dan waktu modifikasi."""

    def __init__(self, parser: RiskWorkbookParser) -> None:
        self.parser = parser
        self._lock = RLock()
        self._signature: tuple[str, int, int] | None = None
        self._payload: dict[str, Any] | None = None

    def set_excel_path(self, excel_path: str | os.PathLike[str]) -> None:
        with self._lock:
            self.parser.excel_path = Path(excel_path)
            self._signature = None
            self._payload = None

    def _current_signature(self) -> tuple[str, int, int]:
        path = self.parser.excel_path
        stat = path.stat()
        return (str(path.resolve()), stat.st_mtime_ns, stat.st_size)

    def get(self, force_refresh: bool = False) -> dict[str, Any]:
        with self._lock:
            signature = self._current_signature()
            if force_refresh or self._payload is None or signature != self._signature:
                self._payload = self.parser.parse()
                self._signature = signature
            return self._payload

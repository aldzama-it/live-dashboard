from __future__ import annotations

import os
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from threading import RLock
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

MONTHS = {
    1: ("Januari", "January", "Jan"),
    2: ("Februari", "February", "Feb"),
    3: ("Maret", "March", "Mar"),
    4: ("April", "Apr"),
    5: ("Mei", "May"),
    6: ("Juni", "June", "Jun"),
    7: ("Juli", "July", "Jul"),
    8: ("Agustus", "August", "Aug"),
    9: ("September", "Sep", "Sept"),
    10: ("Oktober", "October", "Oct"),
    11: ("November", "Nov"),
    12: ("Desember", "December", "Dec"),
}
MONTH_LABELS = {number: names[0] for number, names in MONTHS.items()}

STATUS_LABELS = {
    "done": "Selesai",
    "in_progress": "Sedang berjalan",
    "open": "Belum dimulai",
    "unknown": "Status belum jelas",
}
STATUS_ORDER = ["done", "in_progress", "open", "unknown"]

PRIORITY_ORDER = ["Critical", "High", "Medium", "Low", "Belum ditentukan"]


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\u00a0", " ").replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def percent(part: int, total: int) -> float:
    return round((part / total) * 100, 1) if total else 0.0


def sheet_period(title: str) -> tuple[int, int] | None:
    normalized = normalize_text(title)
    lowered = normalized.casefold()
    year_match = re.search(r"\b(20\d{2})\b", lowered)
    if not year_match:
        return None
    year = int(year_match.group(1))
    for month_number, aliases in MONTHS.items():
        for alias in aliases:
            if re.search(rf"\b{re.escape(alias.casefold())}\b", lowered):
                return year, month_number
    return None


def normalize_status(value: Any) -> str:
    key = normalize_key(value)
    if not key:
        return "unknown"
    if any(token in key for token in ("done", "selesai", "closed", "complete", "completed")):
        return "done"
    if any(token in key for token in (
        "in progress", "on progress", "ongoing", "on going", "progress", "berjalan", "proses",
    )):
        return "in_progress"
    if any(token in key for token in (
        "open", "pending", "not started", "belum", "todo", "to do",
    )):
        return "open"
    return "unknown"


def normalize_priority(value: Any) -> str:
    key = normalize_key(value)
    if "critical" in key or "kritis" in key:
        return "Critical"
    if "high" in key or "tinggi" in key:
        return "High"
    if "medium" in key or "sedang" in key:
        return "Medium"
    if "low" in key or "rendah" in key:
        return "Low"
    return "Belum ditentukan"


def format_due_date(value: Any) -> tuple[str, str | None]:
    if isinstance(value, datetime):
        label = f"{value.day:02d} {MONTH_LABELS[value.month][:3]} {value.year}"
        return label, value.date().isoformat()
    if isinstance(value, date):
        label = f"{value.day:02d} {MONTH_LABELS[value.month][:3]} {value.year}"
        return label, value.isoformat()
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = normalize_text(value)
    return (text or "-"), None


def split_owner_terms(value: str) -> list[str]:
    text = normalize_text(value)
    if not text or text == "-":
        return []
    terms = re.split(r"\s*(?:/|\+|&|,|\band\b|\bdan\b)\s*", text, flags=re.IGNORECASE)
    result: list[str] = []
    seen: set[str] = set()
    for term in terms:
        cleaned = normalize_text(term).strip(" -")
        key = normalize_key(cleaned)
        if not key or key in seen:
            continue
        seen.add(key)
        result.append(cleaned)
    return result


class FollowUpWorkbookParser:
    """Membaca action item rapat BoD per bulan, PIC/divisi, status, dan target."""

    HEADER_ALIASES = {
        "number": {"no", "nomor", "number"},
        "action_item": {
            "action item", "task", "tindakan", "poin tindakan", "poin tindakan utama",
        },
        "pic": {"pic", "owner", "penanggung jawab", "responsible", "person in charge"},
        "due_date": {"planned due date", "due date", "target waktu", "target date", "deadline"},
        "status": {"status", "progress status"},
        "follow_up": {"tindak lanjut", "follow up", "follow-up", "new target", "update"},
        "notes": {"notes", "catatan", "note", "remarks"},
        "priority": {"priority", "risk", "risk level", "prioritas"},
    }

    def __init__(
        self,
        excel_path: str | os.PathLike[str],
        max_scan_rows: int = 3000,
        max_scan_columns: int = 30,
    ) -> None:
        self.excel_path = Path(excel_path)
        self.max_scan_rows = max_scan_rows
        self.max_scan_columns = max_scan_columns

    def _find_header(self, worksheet: Worksheet) -> tuple[int, dict[str, int]]:
        max_row = min(worksheet.max_row, 20)
        max_column = min(worksheet.max_column, self.max_scan_columns)
        best: tuple[int, dict[str, int]] | None = None

        for row_number in range(1, max_row + 1):
            mapped: dict[str, int] = {}
            for column_number in range(1, max_column + 1):
                key = normalize_key(worksheet.cell(row_number, column_number).value)
                if not key:
                    continue
                for field, aliases in self.HEADER_ALIASES.items():
                    if key in aliases or any(key.startswith(f"{alias} ") for alias in aliases):
                        mapped.setdefault(field, column_number)
            if "action_item" in mapped and "pic" in mapped:
                if "status" in mapped:
                    return row_number, mapped
                best = (row_number, mapped)
        if best:
            return best
        raise ValueError(f"Header Action Item dan PIC tidak ditemukan pada sheet {worksheet.title}.")

    @staticmethod
    def _cell(ws: Worksheet, row: int, columns: dict[str, int], field: str) -> Any:
        column = columns.get(field)
        return ws.cell(row, column).value if column else None

    def _parse_sheet(self, worksheet: Worksheet, year: int, month: int) -> list[dict[str, Any]]:
        header_row, columns = self._find_header(worksheet)
        records: list[dict[str, Any]] = []
        max_row = min(worksheet.max_row, self.max_scan_rows)
        month_key = f"{year:04d}-{month:02d}"
        month_label = f"{MONTH_LABELS[month]} {year}"
        today = date.today()

        for row_number in range(header_row + 1, max_row + 1):
            action_item = normalize_text(self._cell(worksheet, row_number, columns, "action_item"))
            if not action_item:
                continue
            if normalize_key(action_item) in {"action item", "task", "tindakan"}:
                continue

            pic = normalize_text(self._cell(worksheet, row_number, columns, "pic")) or "Belum ditentukan"
            status = normalize_status(self._cell(worksheet, row_number, columns, "status"))
            due_label, due_iso = format_due_date(self._cell(worksheet, row_number, columns, "due_date"))
            overdue = bool(
                due_iso
                and status != "done"
                and date.fromisoformat(due_iso) < today
            )
            number_value = normalize_text(self._cell(worksheet, row_number, columns, "number"))
            priority = normalize_priority(self._cell(worksheet, row_number, columns, "priority"))
            follow_up = normalize_text(self._cell(worksheet, row_number, columns, "follow_up")) or "-"
            notes = normalize_text(self._cell(worksheet, row_number, columns, "notes")) or "-"

            records.append({
                "id": f"{month_key}:{row_number}",
                "number": number_value or str(len(records) + 1),
                "month": month_key,
                "month_label": month_label,
                "year": year,
                "month_number": month,
                "action_item": action_item,
                "pic": pic,
                "owner_terms": split_owner_terms(pic),
                "due_date": due_label,
                "due_date_iso": due_iso,
                "status": status,
                "status_label": STATUS_LABELS[status],
                "priority": priority,
                "follow_up": follow_up,
                "notes": notes,
                "overdue": overdue,
                "source_sheet": worksheet.title,
                "source_row": row_number,
            })
        return records

    def parse(self) -> dict[str, Any]:
        if not self.excel_path.is_file():
            raise FileNotFoundError(f"File Follow-up Evaluasi BoD tidak ditemukan: {self.excel_path}")

        workbook = load_workbook(self.excel_path, data_only=True, read_only=False)
        try:
            sheet_periods: list[tuple[Worksheet, int, int]] = []
            for worksheet in workbook.worksheets:
                title_key = normalize_key(worksheet.title)
                if title_key in {"consolidated", "maret dept"} or "dept" in title_key:
                    continue
                period = sheet_period(worksheet.title)
                if period is None:
                    continue
                year, month = period
                if "bod" not in title_key and "direksi" not in title_key:
                    continue
                sheet_periods.append((worksheet, year, month))

            if not sheet_periods:
                raise ValueError("Sheet BoD bulanan tidak ditemukan. Nama sheet perlu memuat bulan dan tahun.")

            records: list[dict[str, Any]] = []
            diagnostics: list[dict[str, Any]] = []
            for worksheet, year, month in sorted(sheet_periods, key=lambda item: (item[1], item[2])):
                try:
                    sheet_records = self._parse_sheet(worksheet, year, month)
                    records.extend(sheet_records)
                    diagnostics.append({
                        "sheet": worksheet.title,
                        "month": f"{year:04d}-{month:02d}",
                        "records": len(sheet_records),
                        "status": "ok",
                    })
                except ValueError as exc:
                    diagnostics.append({
                        "sheet": worksheet.title,
                        "month": f"{year:04d}-{month:02d}",
                        "records": 0,
                        "status": "skipped",
                        "detail": str(exc),
                    })

            if not records:
                raise ValueError("Tidak ada action item BoD yang berhasil dibaca dari workbook.")

            records.sort(key=lambda item: (item["year"], item["month_number"], item["source_row"]))
            status_counts = Counter(item["status"] for item in records)
            priority_counts = Counter(item["priority"] for item in records)
            owner_counts = Counter(item["pic"] for item in records)
            overdue_count = sum(1 for item in records if item["overdue"])
            total = len(records)
            done = status_counts["done"]

            month_rows: list[dict[str, Any]] = []
            for (year, month) in sorted({(item["year"], item["month_number"]) for item in records}):
                month_records = [item for item in records if item["year"] == year and item["month_number"] == month]
                month_status = Counter(item["status"] for item in month_records)
                month_total = len(month_records)
                month_done = month_status["done"]
                month_rows.append({
                    "key": f"{year:04d}-{month:02d}",
                    "label": f"{MONTH_LABELS[month]} {year}",
                    "short_label": f"{MONTH_LABELS[month][:3]} {str(year)[-2:]}",
                    "total": month_total,
                    "done": month_done,
                    "in_progress": month_status["in_progress"],
                    "open": month_status["open"],
                    "unknown": month_status["unknown"],
                    "overdue": sum(1 for item in month_records if item["overdue"]),
                    "completion_percentage": percent(month_done, month_total),
                })

            status_distribution = [
                {
                    "status": status,
                    "label": STATUS_LABELS[status],
                    "count": status_counts[status],
                    "percentage": percent(status_counts[status], total),
                }
                for status in STATUS_ORDER
                if status_counts[status] > 0
            ]

            priority_distribution = [
                {
                    "priority": priority,
                    "count": priority_counts[priority],
                    "percentage": percent(priority_counts[priority], total),
                }
                for priority in PRIORITY_ORDER
                if priority_counts[priority] > 0
            ]

            owner_distribution = [
                {
                    "owner": owner,
                    "count": count,
                    "done": sum(1 for item in records if item["pic"] == owner and item["status"] == "done"),
                    "in_progress": sum(1 for item in records if item["pic"] == owner and item["status"] == "in_progress"),
                    "open": sum(1 for item in records if item["pic"] == owner and item["status"] == "open"),
                }
                for owner, count in sorted(owner_counts.items(), key=lambda item: (-item[1], item[0].casefold()))
            ]

            source_stat = self.excel_path.stat()
            return {
                "available": True,
                "total_tasks": total,
                "summary": {
                    "total": total,
                    "done": done,
                    "in_progress": status_counts["in_progress"],
                    "open": status_counts["open"],
                    "unknown": status_counts["unknown"],
                    "overdue": overdue_count,
                    "completion_percentage": percent(done, total),
                },
                "months": month_rows,
                "latest_month": month_rows[-1]["key"] if month_rows else "",
                "status_distribution": status_distribution,
                "priority_distribution": priority_distribution,
                "owner_distribution": owner_distribution,
                "records": records,
                "filters": {
                    "months": [{"value": item["key"], "label": item["label"], "count": item["total"]} for item in month_rows],
                    "statuses": [{"value": item["status"], "label": item["label"], "count": item["count"]} for item in status_distribution],
                    "owners": [{"value": item["owner"], "label": item["owner"], "count": item["count"]} for item in owner_distribution],
                    "priorities": [{"value": item["priority"], "label": item["priority"], "count": item["count"]} for item in priority_distribution],
                },
                "source": {
                    "file_name": self.excel_path.name,
                    "path": str(self.excel_path),
                    "size_bytes": source_stat.st_size,
                    "modified_at": datetime.fromtimestamp(source_stat.st_mtime).isoformat(timespec="seconds"),
                    "sheet_count": len(sheet_periods),
                },
                "diagnostics": diagnostics,
            }
        finally:
            workbook.close()


class CachedFollowUpRepository:
    def __init__(self, parser: FollowUpWorkbookParser) -> None:
        self.parser = parser
        self._lock = RLock()
        self._cache_key: tuple[str, int, int] | None = None
        self._cache: dict[str, Any] | None = None

    def set_excel_path(self, path: str | os.PathLike[str]) -> None:
        with self._lock:
            self.parser.excel_path = Path(path)
            self._cache_key = None
            self._cache = None

    def get(self, force_refresh: bool = False) -> dict[str, Any]:
        path = self.parser.excel_path
        stat = path.stat()
        cache_key = (str(path.resolve()), stat.st_mtime_ns, stat.st_size)
        with self._lock:
            if not force_refresh and self._cache_key == cache_key and self._cache is not None:
                return self._cache
            parsed = self.parser.parse()
            self._cache_key = cache_key
            self._cache = parsed
            return parsed

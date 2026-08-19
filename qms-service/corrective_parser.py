from __future__ import annotations

import os
import re
import warnings
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from threading import RLock
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


STATUS_ORDER = ["open", "in_progress", "closed", "unknown"]
STATUS_LABELS = {
    "open": "Open",
    "in_progress": "On Progress",
    "closed": "Closed",
    "unknown": "Status belum jelas",
}

INDONESIAN_MONTHS = {
    "januari": 1,
    "jan": 1,
    "februari": 2,
    "feb": 2,
    "maret": 3,
    "mar": 3,
    "april": 4,
    "apr": 4,
    "mei": 5,
    "may": 5,
    "juni": 6,
    "jun": 6,
    "juli": 7,
    "jul": 7,
    "agustus": 8,
    "agu": 8,
    "aug": 8,
    "september": 9,
    "sep": 9,
    "oktober": 10,
    "okt": 10,
    "oct": 10,
    "november": 11,
    "nov": 11,
    "desember": 12,
    "des": 12,
    "dec": 12,
}

MONTH_SHORT = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "Mei",
    6: "Jun",
    7: "Jul",
    8: "Agu",
    9: "Sep",
    10: "Okt",
    11: "Nov",
    12: "Des",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def percentage(part: int, total: int) -> float:
    return round((part / total) * 100, 1) if total else 0.0


def normalize_status(value: Any) -> str:
    key = normalize_key(value)
    if not key:
        return "unknown"
    if any(token in key for token in ("closed", "close", "selesai", "done", "complete", "completed")):
        return "closed"
    if any(token in key for token in (
        "in progress",
        "on progress",
        "ongoing",
        "on going",
        "progress",
        "berjalan",
        "proses",
    )):
        return "in_progress"
    if any(token in key for token in ("open", "pending", "not started", "belum dimulai", "todo", "to do")):
        return "open"
    return "unknown"


def parse_date_text(value: str) -> date | None:
    text = normalize_text(value).casefold().replace(".", " ").replace(",", " ")
    if not text:
        return None

    numeric_match = re.search(r"\b(\d{1,2})[/-](\d{1,2})[/-](20\d{2}|\d{2})\b", text)
    if numeric_match:
        day, month, year = map(int, numeric_match.groups())
        if year < 100:
            year += 2000
        try:
            return date(year, month, day)
        except ValueError:
            return None

    word_match = re.search(r"\b(\d{1,2})\s+([a-z]+)\s+(20\d{2}|\d{2})\b", text)
    if word_match:
        day = int(word_match.group(1))
        month = INDONESIAN_MONTHS.get(word_match.group(2))
        year = int(word_match.group(3))
        if year < 100:
            year += 2000
        if month:
            try:
                return date(year, month, day)
            except ValueError:
                return None
    return None


def format_date_value(value: Any) -> tuple[str, str | None]:
    parsed: date | None = None
    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        text = normalize_text(value)
        parsed = parse_date_text(text)
        if parsed is None:
            return (text or "-"), None

    return f"{parsed.day:02d} {MONTH_SHORT[parsed.month]} {parsed.year}", parsed.isoformat()


def plain_value(value: Any) -> str:
    if value is None:
        return "-"
    if isinstance(value, bool):
        return "Ya" if value else "Tidak"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = normalize_text(value)
    return text or "-"


class CorrectiveActionWorkbookParser:
    """Membaca rekap status dan detail dari workbook Corrective Action Register."""

    HEADER_ALIASES = {
        "car_number": {"no car", "nomor car", "car no", "car number", "no corrective action"},
        "audit_date": {"tanggal audit", "audit date"},
        "auditor": {"auditor"},
        "audit_type": {"audit type", "jenis audit"},
        "clause": {"klausul iso", "iso clause", "clause"},
        "division": {"divisi terkait", "division", "divisi"},
        "department": {"departement", "department", "departemen"},
        "description": {"deskripsi ketidaksesuaian", "description", "finding description", "temuan"},
        "grade": {"grading nc", "nc grade", "grade", "grading"},
        "pic": {"pic penanggung jawab", "pic", "penanggung jawab", "owner"},
        "target_date": {"target selesai", "target date", "due date", "deadline"},
        "root_cause": {"root cause analysis", "root cause", "akar masalah"},
        "corrective_action": {"tindakan perbaikan", "corrective action", "action"},
        "preventive_action": {"tindakan pencegahan", "preventive action"},
        "evaluation": {"evaluasi", "evaluation"},
        "suitability": {"suitability", "kesesuaian"},
        "verifier": {"verifikasi oleh", "verified by", "verifier"},
        "verification_date": {"tanggal verifikasi", "verification date"},
        "status": {"status", "car status", "progress status"},
        "auditor_note": {"catatan auditor", "auditor note"},
        "qms_note": {"catatan qms", "qms note"},
    }

    def __init__(
        self,
        excel_path: str | os.PathLike[str],
        sheet_name: str = "CAR Register",
        max_scan_rows: int = 5000,
        max_scan_columns: int = 40,
    ) -> None:
        self.excel_path = Path(excel_path)
        self.sheet_name = sheet_name
        self.max_scan_rows = max_scan_rows
        self.max_scan_columns = max_scan_columns

    def _find_sheet(self, workbook) -> Worksheet:
        expected = normalize_key(self.sheet_name)
        for worksheet in workbook.worksheets:
            if normalize_key(worksheet.title) == expected:
                return worksheet
        for worksheet in workbook.worksheets:
            key = normalize_key(worksheet.title)
            if "car" in key and "register" in key:
                return worksheet
        raise ValueError("Sheet CAR Register tidak ditemukan pada workbook.")

    def _find_header(self, worksheet: Worksheet) -> tuple[int, dict[str, int]]:
        max_row = min(worksheet.max_row, 30)
        max_column = min(worksheet.max_column, self.max_scan_columns)
        best: tuple[int, dict[str, int], int] | None = None

        for row_number in range(1, max_row + 1):
            mapped: dict[str, int] = {}
            for column_number in range(1, max_column + 1):
                key = normalize_key(worksheet.cell(row_number, column_number).value)
                if not key:
                    continue
                for field, aliases in self.HEADER_ALIASES.items():
                    if key in aliases or any(key.startswith(f"{alias} ") for alias in aliases):
                        mapped.setdefault(field, column_number)
            score = len(mapped)
            if {"car_number", "status"}.issubset(mapped):
                return row_number, mapped
            if score >= 4 and (best is None or score > best[2]):
                best = (row_number, mapped, score)

        if best and {"car_number", "status"}.issubset(best[1]):
            return best[0], best[1]
        raise ValueError("Header NO CAR dan Status tidak ditemukan pada sheet CAR Register.")

    @staticmethod
    def _cell(worksheet: Worksheet, row: int, columns: dict[str, int], field: str) -> Any:
        column = columns.get(field)
        return worksheet.cell(row, column).value if column else None

    def parse(self) -> dict[str, Any]:
        if not self.excel_path.is_file():
            raise FileNotFoundError(f"File Corrective Action Register tidak ditemukan: {self.excel_path}")

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
            workbook = load_workbook(self.excel_path, data_only=True, read_only=False)

        try:
            worksheet = self._find_sheet(workbook)
            header_row, columns = self._find_header(worksheet)
            max_row = min(worksheet.max_row, self.max_scan_rows)
            today = date.today()
            records: list[dict[str, Any]] = []
            blank_streak = 0

            for row_number in range(header_row + 1, max_row + 1):
                car_number = normalize_text(self._cell(worksheet, row_number, columns, "car_number"))
                if not car_number:
                    if records:
                        blank_streak += 1
                        if blank_streak >= 100:
                            break
                    continue
                blank_streak = 0

                status = normalize_status(self._cell(worksheet, row_number, columns, "status"))
                audit_date, audit_date_iso = format_date_value(
                    self._cell(worksheet, row_number, columns, "audit_date")
                )
                target_date, target_date_iso = format_date_value(
                    self._cell(worksheet, row_number, columns, "target_date")
                )
                verification_date, verification_date_iso = format_date_value(
                    self._cell(worksheet, row_number, columns, "verification_date")
                )
                overdue = bool(
                    target_date_iso
                    and status != "closed"
                    and date.fromisoformat(target_date_iso) < today
                )

                records.append({
                    "id": f"{worksheet.title}:{row_number}",
                    "car_number": car_number,
                    "audit_date": audit_date,
                    "audit_date_iso": audit_date_iso,
                    "auditor": plain_value(self._cell(worksheet, row_number, columns, "auditor")),
                    "audit_type": plain_value(self._cell(worksheet, row_number, columns, "audit_type")),
                    "clause": plain_value(self._cell(worksheet, row_number, columns, "clause")),
                    "division": plain_value(self._cell(worksheet, row_number, columns, "division")),
                    "department": plain_value(self._cell(worksheet, row_number, columns, "department")),
                    "description": plain_value(self._cell(worksheet, row_number, columns, "description")),
                    "grade": plain_value(self._cell(worksheet, row_number, columns, "grade")),
                    "pic": plain_value(self._cell(worksheet, row_number, columns, "pic")),
                    "target_date": target_date,
                    "target_date_iso": target_date_iso,
                    "root_cause": plain_value(self._cell(worksheet, row_number, columns, "root_cause")),
                    "corrective_action": plain_value(self._cell(worksheet, row_number, columns, "corrective_action")),
                    "preventive_action": plain_value(self._cell(worksheet, row_number, columns, "preventive_action")),
                    "evaluation": plain_value(self._cell(worksheet, row_number, columns, "evaluation")),
                    "suitability": plain_value(self._cell(worksheet, row_number, columns, "suitability")),
                    "verifier": plain_value(self._cell(worksheet, row_number, columns, "verifier")),
                    "verification_date": verification_date,
                    "verification_date_iso": verification_date_iso,
                    "status": status,
                    "status_label": STATUS_LABELS[status],
                    "auditor_note": plain_value(self._cell(worksheet, row_number, columns, "auditor_note")),
                    "qms_note": plain_value(self._cell(worksheet, row_number, columns, "qms_note")),
                    "overdue": overdue,
                    "source_sheet": worksheet.title,
                    "source_row": row_number,
                })

            if not records:
                raise ValueError("Tidak ada data Corrective Action yang berhasil dibaca.")

            status_counts = Counter(item["status"] for item in records)
            division_counts = Counter(item["division"] for item in records)
            audit_type_counts = Counter(item["audit_type"] for item in records)
            grade_counts = Counter(item["grade"] for item in records)
            total = len(records)
            closed = status_counts["closed"]
            overdue_count = sum(1 for item in records if item["overdue"])

            status_distribution = [
                {
                    "status": status,
                    "label": STATUS_LABELS[status],
                    "count": int(status_counts[status]),
                    "percentage": percentage(int(status_counts[status]), total),
                }
                for status in STATUS_ORDER
                if status_counts[status] > 0 or status != "unknown"
            ]

            division_distribution = [
                {
                    "division": division,
                    "count": int(count),
                    "open": sum(1 for item in records if item["division"] == division and item["status"] == "open"),
                    "in_progress": sum(1 for item in records if item["division"] == division and item["status"] == "in_progress"),
                    "closed": sum(1 for item in records if item["division"] == division and item["status"] == "closed"),
                }
                for division, count in sorted(
                    division_counts.items(),
                    key=lambda item: (-item[1], item[0].casefold()),
                )
            ]

            source_stat = self.excel_path.stat()
            return {
                "available": True,
                "total_actions": total,
                "summary": {
                    "total": total,
                    "open": int(status_counts["open"]),
                    "in_progress": int(status_counts["in_progress"]),
                    "closed": int(closed),
                    "unknown": int(status_counts["unknown"]),
                    "overdue": int(overdue_count),
                    "completion_percentage": percentage(int(closed), total),
                    "active_percentage": percentage(
                        int(status_counts["open"] + status_counts["in_progress"]), total
                    ),
                },
                "status_distribution": status_distribution,
                "division_distribution": division_distribution,
                "records": records,
                "filters": {
                    "statuses": [
                        {"value": item["status"], "label": item["label"], "count": item["count"]}
                        for item in status_distribution
                    ],
                    "divisions": [
                        {"value": item["division"], "label": item["division"], "count": item["count"]}
                        for item in division_distribution
                    ],
                    "audit_types": [
                        {"value": label, "label": label, "count": int(count)}
                        for label, count in sorted(audit_type_counts.items(), key=lambda item: (-item[1], item[0].casefold()))
                    ],
                    "grades": [
                        {"value": label, "label": label, "count": int(count)}
                        for label, count in sorted(grade_counts.items(), key=lambda item: (-item[1], item[0].casefold()))
                    ],
                },
                "source": {
                    "file_name": self.excel_path.name,
                    "path": str(self.excel_path),
                    "sheet_name": worksheet.title,
                    "size_bytes": source_stat.st_size,
                    "modified_at": datetime.fromtimestamp(source_stat.st_mtime).isoformat(timespec="seconds"),
                    "rows_read": total,
                    "header_row": header_row,
                },
            }
        finally:
            workbook.close()


class CachedCorrectiveActionRepository:
    """Cache berdasarkan path, ukuran file, dan waktu modifikasi."""

    def __init__(self, parser: CorrectiveActionWorkbookParser) -> None:
        self.parser = parser
        self._lock = RLock()
        self._signature: tuple[str, int, int] | None = None
        self._payload: dict[str, Any] | None = None

    def set_excel_path(self, excel_path: str | os.PathLike[str]) -> None:
        with self._lock:
            self.parser.excel_path = Path(excel_path)
            self._signature = None
            self._payload = None

    def _current_signature(self) -> tuple[str, int, int]:
        path = self.parser.excel_path
        stat = path.stat()
        return (str(path.resolve()), stat.st_mtime_ns, stat.st_size)

    def get(self, force_refresh: bool = False) -> dict[str, Any]:
        with self._lock:
            signature = self._current_signature()
            if force_refresh or self._payload is None or signature != self._signature:
                self._payload = self.parser.parse()
                self._signature = signature
            return self._payload

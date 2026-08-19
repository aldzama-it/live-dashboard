from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path
from threading import RLock
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.worksheet.worksheet import Worksheet


def normalize_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.replace("\u00a0", " ").replace("\r", " ").replace("\n", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value: Any) -> str:
    text = normalize_text(value).casefold()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def as_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = normalize_text(value)
    if not text:
        return None
    text = text.replace("%", "").replace(" ", "")
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def average(values: Iterable[Any]) -> float | None:
    numbers = [number for value in values if (number := as_number(value)) is not None]
    if not numbers:
        return None
    return sum(numbers) / len(numbers)


class KaizenWorkbookParser:
    """Membaca ranking Kaizen dan menghitung Top 10 secara otomatis.

    Ranking utama menggunakan Weighted Average 60/40:
      60% Director's Average + 40% Manager's Average.

    Jika workbook menyimpan rumus AVERAGE untuk Director/Manager, parser membaca
    rentang penilai dari rumus tersebut lalu menghitung ulang nilai rata-ratanya
    dari kolom penilai. Dengan begitu ranking tidak bergantung pada sheet Winner.
    """

    NAME_ALIASES = {
        "nama",
        "name",
        "participant name",
        "participant",
        "participant code name",
    }
    IDEA_NUMBER_ALIASES = {"no", "no ide", "idea no", "idea number", "nomor ide"}
    AREA_ALIASES = {"area kerja", "work area", "area"}
    IDEA_ALIASES = {"ide kaizen", "idea kaizen", "kaizen idea", "ide"}
    DIRECTOR_AVG_ALIASES = {"director s average", "directors average", "director average"}
    MANAGER_AVG_ALIASES = {"manager s average", "managers average", "manager average"}
    WEIGHTED_ALIASES = {
        "weighed average 60 40",
        "weighted average 60 40",
        "weighted 60 40",
        "weighed 60 40",
    }

    def __init__(
        self,
        excel_path: str | os.PathLike[str],
        top_limit: int = 10,
        max_scan_rows: int = 5000,
        max_scan_columns: int = 80,
    ) -> None:
        self.excel_path = Path(excel_path)
        self.top_limit = max(1, min(int(top_limit), 50))
        self.max_scan_rows = max_scan_rows
        self.max_scan_columns = max_scan_columns

    @staticmethod
    def _find_sheet(workbook) -> Worksheet:
        for worksheet in workbook.worksheets:
            key = normalize_key(worksheet.title)
            if key == "consolidated scores":
                return worksheet
        for worksheet in workbook.worksheets:
            key = normalize_key(worksheet.title)
            if "consolidated" in key and "score" in key:
                return worksheet
        raise ValueError("Sheet 'Consolidated Scores' tidak ditemukan pada file Kaizen.")

    def _find_header(self, worksheet: Worksheet) -> tuple[int, dict[str, int]]:
        max_row = min(worksheet.max_row, 15)
        max_column = min(worksheet.max_column, self.max_scan_columns)

        for row_number in range(1, max_row + 1):
            columns: dict[str, int] = {}
            for column_number in range(1, max_column + 1):
                key = normalize_key(worksheet.cell(row_number, column_number).value)
                if not key:
                    continue
                if key in self.NAME_ALIASES:
                    columns.setdefault("name", column_number)
                if key in self.IDEA_NUMBER_ALIASES:
                    columns.setdefault("idea_number", column_number)
                if key in self.AREA_ALIASES:
                    columns.setdefault("area", column_number)
                if key in self.IDEA_ALIASES:
                    columns.setdefault("idea", column_number)
                if key in self.DIRECTOR_AVG_ALIASES:
                    columns.setdefault("director_average", column_number)
                if key in self.MANAGER_AVG_ALIASES:
                    columns.setdefault("manager_average", column_number)
                if key in self.WEIGHTED_ALIASES:
                    columns.setdefault("weighted", column_number)

            if "name" in columns and (
                "weighted" in columns
                or ("director_average" in columns and "manager_average" in columns)
            ):
                return row_number, columns

        raise ValueError(
            "Header NAMA dan kolom Weighted Average 60/40 atau Director/Manager Average tidak ditemukan."
        )

    @staticmethod
    def _average_range_from_formula(formula: Any, expected_row: int) -> tuple[int, int] | None:
        if not isinstance(formula, str) or not formula.startswith("="):
            return None
        match = re.search(
            r"AVERAGE\s*\(\s*\$?([A-Z]{1,3})\$?(\d+)\s*:\s*\$?([A-Z]{1,3})\$?(\d+)\s*\)",
            formula,
            flags=re.IGNORECASE,
        )
        if not match:
            return None
        start_col, start_row, end_col, end_row = match.groups()
        if int(start_row) != expected_row or int(end_row) != expected_row:
            return None
        min_col, _, max_col, _ = range_boundaries(f"{start_col}{expected_row}:{end_col}{expected_row}")
        return min_col, max_col

    @staticmethod
    def _row_average_from_formula(
        formula_ws: Worksheet,
        values_ws: Worksheet,
        row_number: int,
        average_column: int | None,
    ) -> float | None:
        if not average_column:
            return None
        formula = formula_ws.cell(row_number, average_column).value
        bounds = KaizenWorkbookParser._average_range_from_formula(formula, row_number)
        if bounds is None:
            return as_number(values_ws.cell(row_number, average_column).value)
        start_col, end_col = bounds
        return average(values_ws.cell(row_number, column).value for column in range(start_col, end_col + 1))

    def parse(self) -> dict[str, Any]:
        if not self.excel_path.is_file():
            raise FileNotFoundError(f"File Kaizen Recap tidak ditemukan: {self.excel_path}")

        # Workbook formula dipakai untuk membaca pola AVERAGE penilai.
        # Workbook data_only dipakai untuk nilai angka hasil perhitungan Excel.
        formula_book = load_workbook(self.excel_path, data_only=False, read_only=False)
        value_book = load_workbook(self.excel_path, data_only=True, read_only=False)

        try:
            formula_ws = self._find_sheet(formula_book)
            value_ws = value_book[formula_ws.title]
            header_row, columns = self._find_header(formula_ws)
            max_row = min(formula_ws.max_row, self.max_scan_rows)

            records: list[dict[str, Any]] = []
            skipped_without_score = 0

            for row_number in range(header_row + 1, max_row + 1):
                name = normalize_text(value_ws.cell(row_number, columns["name"]).value)
                if not name:
                    continue

                director_average = self._row_average_from_formula(
                    formula_ws,
                    value_ws,
                    row_number,
                    columns.get("director_average"),
                )
                manager_average = self._row_average_from_formula(
                    formula_ws,
                    value_ws,
                    row_number,
                    columns.get("manager_average"),
                )

                score: float | None = None
                if director_average is not None and manager_average is not None:
                    score = (0.60 * director_average) + (0.40 * manager_average)
                elif columns.get("weighted"):
                    score = as_number(value_ws.cell(row_number, columns["weighted"]).value)

                if score is None:
                    skipped_without_score += 1
                    continue

                idea_number = ""
                if columns.get("idea_number"):
                    idea_number = normalize_text(value_ws.cell(row_number, columns["idea_number"]).value)
                area = ""
                if columns.get("area"):
                    area = normalize_text(value_ws.cell(row_number, columns["area"]).value)
                idea = ""
                if columns.get("idea"):
                    idea = normalize_text(value_ws.cell(row_number, columns["idea"]).value)

                records.append(
                    {
                        "name": name,
                        "score_raw": float(score),
                        "director_average": None if director_average is None else round(director_average, 6),
                        "manager_average": None if manager_average is None else round(manager_average, 6),
                        "idea_number": idea_number,
                        "area": area,
                        "idea": idea,
                        "source_row": row_number,
                    }
                )

            if not records:
                raise ValueError("Tidak ada peserta Kaizen dengan score yang berhasil dibaca.")

            records.sort(
                key=lambda item: (
                    -item["score_raw"],
                    item["name"].casefold(),
                    item["source_row"],
                )
            )

            all_rows: list[dict[str, Any]] = []
            for index, item in enumerate(records, start=1):
                all_rows.append(
                    {
                        "id": f"row:{item['source_row']}",
                        "rank": index,
                        "score": round(item["score_raw"], 2),
                        "name": item["name"],
                        "idea_number": item["idea_number"],
                        "area": item["area"],
                        "idea": item["idea"],
                        "director_average": None if item["director_average"] is None else round(item["director_average"], 2),
                        "manager_average": None if item["manager_average"] is None else round(item["manager_average"], 2),
                        "source_row": item["source_row"],
                    }
                )

            top_rows = [dict(item) for item in all_rows[: self.top_limit]]

            source_stat = self.excel_path.stat()
            return {
                "available": True,
                "ranking_basis": "Weighted Average 60/40",
                "ranking_note": "Score = 60% Director's Average + 40% Manager's Average",
                "visible_limit": 5,
                "top_limit": self.top_limit,
                "total_scored_entries": len(records),
                "skipped_without_score": skipped_without_score,
                "winner": top_rows[0] if top_rows else None,
                "top_10": top_rows,
                "records": all_rows,
                "source": {
                    "file_name": self.excel_path.name,
                    "path": str(self.excel_path),
                    "size_bytes": source_stat.st_size,
                    "modified_at": datetime.fromtimestamp(source_stat.st_mtime).isoformat(timespec="seconds"),
                    "sheet_name": formula_ws.title,
                    "header_row": header_row,
                },
            }
        finally:
            formula_book.close()
            value_book.close()


class CachedKaizenRepository:
    def __init__(self, parser: KaizenWorkbookParser) -> None:
        self.parser = parser
        self._lock = RLock()
        self._cache_key: tuple[str, int, int] | None = None
        self._cache: dict[str, Any] | None = None

    def set_excel_path(self, path: str | os.PathLike[str]) -> None:
        with self._lock:
            self.parser.excel_path = Path(path)
            self._cache_key = None
            self._cache = None

    def get(self, force_refresh: bool = False) -> dict[str, Any]:
        path = self.parser.excel_path
        stat = path.stat()
        cache_key = (str(path.resolve()), stat.st_mtime_ns, stat.st_size)
        with self._lock:
            if not force_refresh and self._cache_key == cache_key and self._cache is not None:
                return self._cache
            parsed = self.parser.parse()
            self._cache_key = cache_key
            self._cache = parsed
            return parsed
